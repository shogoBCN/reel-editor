"""
Build Angélica's Excel brief (Brief + Instrucciones).

She fills three fields at the top and one row per overlay (times, photo,
natural-language «Qué quieres»). Cursor turns that into brief.yaml.
Run from repo root:

    python scripts/generate_template_xlsx.py
    python scripts/generate_template_xlsx.py \\
        --csv-dir projects/<slug>/brief_sheet \\
        --overlays projects/<slug>/overlays \\
        --out projects/<slug>/BRIEF_Angelica.xlsx
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

REPOSITORY_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPOSITORY_ROOT))

from modules.video.timing import (
    format_seconds_as_sheet_timestamp,
    parse_timestamp_to_seconds,
)

HEADER_ROW = 6
FIRST_DATA_ROW = 7
BLANK_DATA_ROWS = 12
INSTRUCTIONS_PATH = REPOSITORY_ROOT / "templates" / "angelica_brief" / "00_instrucciones.txt"
TEXT_NUMBER_FORMAT = "@"

HINT = (
    "No recortes el video: marca dónde empieza el reel y cuándo terminas de hablar. "
    "Tiempos: 1:29,5 (minutos:segundos, décimos con coma). Reloj = TU grabación. "
    "Pega la foto del tamaño que sea. En «Qué quieres»: qué se ve, "
    "dónde (derecha / izquierda / arriba) y tamaño (grande / mediano / chico). "
    "Dos fotos juntas = dos filas. Ver pestaña Instrucciones."
)


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    """Load a UTF-8 CSV as dicts (BOM-safe)."""
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _project_fields_from_csv(csv_directory: Path) -> dict[str, str]:
    """Read título / cortar_inicio / fin_de_habla from 01_proyecto.csv."""
    path = csv_directory / "01_proyecto.csv"
    fields = {"titulo": "", "cortar_inicio": "", "fin_de_habla": ""}
    if not path.is_file():
        return fields
    for row in _read_csv_rows(path):
        key = (row.get("clave") or "").strip().lower()
        value = (row.get("valor") or "").strip()
        if key == "titulo":
            fields["titulo"] = value
        elif key in ("cortar_inicio", "empieza_el_reel", "trim_start"):
            fields["cortar_inicio"] = _sheet_time_cell(value)
        elif key == "fin_de_habla":
            try:
                fields["fin_de_habla"] = format_seconds_as_sheet_timestamp(
                    parse_timestamp_to_seconds(value)
                )
            except ValueError:
                fields["fin_de_habla"] = value
    return fields


def _sheet_time_cell(raw: str) -> str:
    """Show a CSV timestamp as Angélica writes it (``1:29,5``)."""
    text = (raw or "").strip()
    if not text:
        return ""
    try:
        return format_seconds_as_sheet_timestamp(parse_timestamp_to_seconds(text))
    except ValueError:
        return text


def _instruction_blocks() -> list[tuple[str, str]]:
    """Load heading/body pairs for the Instrucciones tab.

    Returns:
        ``("heading"|"body", text)`` rows. Skips the markdown ``#`` title.
    """
    if not INSTRUCTIONS_PATH.is_file():
        return [("body", HINT)]
    raw = INSTRUCTIONS_PATH.read_text(encoding="utf-8").strip()
    blocks: list[tuple[str, str]] = []
    for chunk in raw.split("\n\n"):
        lines = [line.rstrip() for line in chunk.splitlines() if line.strip()]
        if not lines:
            continue
        if lines[0].startswith("# ") and not blocks:
            continue
        if lines[0].startswith("## "):
            blocks.append(("heading", lines[0][3:].strip()))
            lines = lines[1:]
            if not lines:
                continue
        cleaned = []
        for line in lines:
            if line.startswith("- "):
                cleaned.append("• " + line[2:].strip())
            else:
                cleaned.append(line.strip())
        blocks.append(("body", "\n".join(cleaned)))
    return blocks or [("body", HINT)]


def _overlay_rows_from_csv(csv_directory: Path) -> list[dict[str, str]]:
    """Flatten the detailed overlay CSV into Desde / Hasta / Foto / Qué quieres."""
    path = csv_directory / "02_imagenes_y_tiempos.csv"
    if not path.is_file():
        return []
    simple: list[dict[str, str]] = []
    for row in _read_csv_rows(path):
        kind = (row.get("tipo") or "").strip().lower()
        side = (row.get("lado") or "").strip().lower()
        if side in ("gancho", "centro", "hook"):
            side = "arriba"
        text = (row.get("texto") or "").strip()
        file_name = (row.get("archivo_imagen") or "").strip()
        if kind == "enfoque":
            text = text or "ENFOQUE INTEGRAL"
            file_name = ""
        if kind == "etiqueta":
            file_name = ""
        nota = (row.get("que_es") or "").strip()
        extra = (row.get("notas_edicion") or "").strip()
        lado = side
        # Filled examples store the human «Qué quieres» sentence in notas_edicion.
        if extra:
            want = extra
        elif kind == "etiqueta" and text:
            want = f"{text} arriba en el cielo"
        elif kind == "enfoque":
            want = text or "ENFOQUE INTEGRAL arriba en el cielo (sin foto, se genera)"
        elif lado:
            want = nota if nota else f"{lado} de la cabeza"
            if lado and want and lado not in want.lower():
                want = f"{want}. {lado} de la cabeza."
        else:
            want = nota
        simple.append(
            {
                "desde": _sheet_time_cell((row.get("tiempo_inicio") or "").strip()),
                "hasta": _sheet_time_cell((row.get("tiempo_fin") or "").strip()),
                "archivo": file_name,
                "quieres": want.strip(),
            }
        )
    return simple


def generate_xlsx(
    destination: Path,
    csv_directory: Path | None = None,
    overlay_directory: Path | None = None,
) -> Path:
    """Write Angélica's workbook: Brief tab plus Instrucciones.

    Args:
        destination: Output ``.xlsx``.
        csv_directory: Optional filled ``01_`` / ``02_`` CSVs to pre-populate.
        overlay_directory: PNGs to embed in the Foto column.

    Returns:
        Path written.

    Raises:
        ImportError: openpyxl is missing.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("pip install openpyxl") from exc

    fields = (
        _project_fields_from_csv(csv_directory)
        if csv_directory
        else {"titulo": "", "cortar_inicio": "", "fin_de_habla": ""}
    )
    overlay_rows = _overlay_rows_from_csv(csv_directory) if csv_directory else []

    header_fill = PatternFill("solid", fgColor="068A93")
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True, color="068A93")
    hint_font = Font(italic=True, color="666666", size=11)
    wrap = Alignment(wrap_text=True, vertical="center")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Brief"

    sheet["A1"] = "Título"
    sheet["B1"] = fields["titulo"]
    sheet["A2"] = "El reel empieza en (reloj de TU video)"
    sheet["B2"] = fields["cortar_inicio"]
    sheet["B2"].number_format = TEXT_NUMBER_FORMAT
    sheet["A3"] = "Terminas de hablar (reloj de TU video)"
    sheet["B3"] = fields["fin_de_habla"]
    sheet["B3"].number_format = TEXT_NUMBER_FORMAT
    for row_index in (1, 2, 3):
        sheet.cell(row_index, 1).font = label_font
        sheet.merge_cells(
            start_row=row_index, start_column=2, end_row=row_index, end_column=4
        )

    sheet.merge_cells("A4:D4")
    sheet["A4"] = HINT
    sheet["A4"].font = hint_font
    sheet["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[4].height = 72

    headers = ["Desde", "Hasta", "Foto", "Qué quieres"]
    for column_index, title in enumerate(headers, start=1):
        cell = sheet.cell(HEADER_ROW, column_index, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    widths = (44, 14, 14, 56)
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    data_count = max(len(overlay_rows), BLANK_DATA_ROWS)
    for offset in range(data_count):
        excel_row = FIRST_DATA_ROW + offset
        sheet.row_dimensions[excel_row].height = 88
        for column_index in (1, 2):
            sheet.cell(excel_row, column_index).number_format = TEXT_NUMBER_FORMAT
            sheet.cell(excel_row, column_index).alignment = wrap
        sheet.cell(excel_row, 4).alignment = wrap
        if offset < len(overlay_rows):
            item = overlay_rows[offset]
            sheet.cell(excel_row, 1, item["desde"])
            sheet.cell(excel_row, 2, item["hasta"])
            sheet.cell(excel_row, 1).number_format = TEXT_NUMBER_FORMAT
            sheet.cell(excel_row, 2).number_format = TEXT_NUMBER_FORMAT
            sheet.cell(excel_row, 4, item["quieres"])
            png_name = item.get("archivo") or ""
            png_path = (
                overlay_directory / png_name
                if overlay_directory and png_name
                else None
            )
            if png_path is not None and png_path.is_file():
                excel_image = ExcelImage(str(png_path))
                excel_image.width = 72
                excel_image.height = 72
                sheet.add_image(excel_image, f"C{excel_row}")

    sheet.freeze_panes = "A7"
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[2].height = 22
    sheet.row_dimensions[3].height = 22
    sheet.row_dimensions[HEADER_ROW].height = 22

    _write_instructions_sheet(workbook)
    workbook.active = 0

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _write_instructions_sheet(workbook) -> None:
    """Add the Instrucciones tab from ``00_instrucciones.txt``."""
    from openpyxl.styles import Alignment, Font, PatternFill

    sheet = workbook.create_sheet("Instrucciones")
    header_fill = PatternFill("solid", fgColor="068A93")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    heading_font = Font(bold=True, color="068A93", size=12)
    body_font = Font(size=12, color="333333")
    wrap = Alignment(wrap_text=True, vertical="top")

    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Cómo llenar el Brief"
    sheet["A1"].fill = header_fill
    sheet["A1"].font = header_font
    sheet["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.column_dimensions["A"].width = 88
    sheet.column_dimensions["B"].width = 12

    excel_row = 3
    for kind, text in _instruction_blocks():
        sheet.merge_cells(
            start_row=excel_row, start_column=1, end_row=excel_row, end_column=2
        )
        cell = sheet.cell(excel_row, 1, text)
        cell.alignment = wrap
        cell.font = heading_font if kind == "heading" else body_font
        line_count = text.count("\n") + 1
        sheet.row_dimensions[excel_row].height = max(22, 18 * line_count + 8)
        excel_row += 1 if kind == "heading" else 2



def main() -> None:
    """Write the blank template, or a filled brief from CSV + PNGs."""
    parser = argparse.ArgumentParser(description="Build Angélica's Excel brief (Brief + Instrucciones)")
    parser.add_argument("--csv-dir", help="Folder with 01_proyecto.csv and 02_imagenes_y_tiempos.csv")
    parser.add_argument("--overlays", help="PNG folder to embed in Foto")
    parser.add_argument(
        "--out",
        help="Destination xlsx (default: templates/angelica_brief/plantilla_reel_angelica.xlsx)",
    )
    args = parser.parse_args()
    destination = (
        Path(args.out)
        if args.out
        else REPOSITORY_ROOT / "templates" / "angelica_brief" / "plantilla_reel_angelica.xlsx"
    )
    path = generate_xlsx(
        destination,
        csv_directory=Path(args.csv_dir) if args.csv_dir else None,
        overlay_directory=Path(args.overlays) if args.overlays else None,
    )
    print("wrote", path)


if __name__ == "__main__":
    main()
