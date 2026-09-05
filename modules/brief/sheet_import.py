"""
Convert Angélica's spreadsheet (CSV folder or xlsx) into a first-pass ``brief.yaml``.

The one-tab Excel (Desde / Hasta / Foto / Qué quieres) is what she fills, in
her own words. Import is a sketch: Cursor amends it to engine-ready YAML
(kinds, placements) using ``.cursor/skills/angelica-reel-brief``. Size words
(grande / mediano / chico) become ``max_w`` / ``max_h``. Legacy multi-tab
CSV/xlsx still works.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from modules.brief.brief_loader import write_brief_yaml
from modules.brief.sticker_size import sticker_box_for_notes
from modules.video.timing import parse_timestamp_to_seconds


# Spanish sheet values → engine placement names.
PLACEMENT_BY_ALIAS = {
    "derecha": "right",
    "right": "right",
    "izquierda": "left",
    "left": "left",
    "centro": "hook_center",
    "center": "hook_center",
    "gancho": "hook_center",
    "arriba": "hook_center",
    "cielo": "hook_center",
    "hook": "hook_center",
    "hook_center": "hook_center",
}

KIND_BY_ALIAS = {
    "sticker": "sticker",
    "imagen": "sticker",
    "png": "sticker",
    "etiqueta": "brush_label",
    "brush_label": "brush_label",
    "texto": "brush_label",
    "enfoque": "enfoque_lockup",
    "enfoque_lockup": "enfoque_lockup",
}


def normalise_column_header(name: str) -> str:
    """Lowercase, unaccent, and snake_case a spreadsheet header.

    Args:
        name: Raw CSV/xlsx header (may include ``á`` / spaces).

    Returns:
        Stable key such as ``tiempo_inicio``.
    """
    cleaned = (
        str(name or "")
        .strip()
        .lower()
        .replace("«", "")
        .replace("»", "")
        .replace("¿", "")
        .replace("?", "")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
        .replace(" ", "_")
    )
    return cleaned.strip("_")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    """Read a UTF-8 CSV (BOM-safe) into dicts keyed by normalised headers.

    Args:
        path: CSV file.

    Returns:
        Non-empty rows only (blank spreadsheet lines are skipped).
    """
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for raw in reader:
            row = {normalise_column_header(key): (value or "").strip() for key, value in raw.items() if key}
            if not any(row.values()):
                continue
            rows.append(row)
        return rows


def read_key_value_csv(path: Path) -> dict[str, str]:
    """Read the Proyecto tab (``clave`` / ``valor`` columns).

    Args:
        path: ``01_proyecto.csv``.

    Returns:
        Normalised key → value.
    """
    mapping: dict[str, str] = {}
    for row in read_csv_rows(path):
        key = row.get("clave") or row.get("key") or row.get("campo")
        value = row.get("valor") or row.get("value") or row.get("dato")
        if key:
            mapping[normalise_column_header(key)] = value or ""
    return mapping


def find_first_existing_file(folder: Path, names: list[str]) -> Path | None:
    """Return the first filename that exists in ``folder``.

    Args:
        folder: Directory to search.
        names: Candidate filenames in preference order.

    Returns:
        Path if found, else None.
    """
    for name in names:
        path = folder / name
        if path.is_file():
            return path
    return None


def map_placement_name(raw: str) -> str:
    """Translate Spanish ``lado`` values to engine placement names.

    Args:
        raw: Cell value (``derecha``, ``izquierda``, ``gancho``, …).

    Returns:
        ``right``, ``left``, or ``hook_center``.

    Raises:
        ValueError: Unknown value — fail closed so a typo does not silently
            default to the right slot.
    """
    key = normalise_column_header(raw)
    if key not in PLACEMENT_BY_ALIAS:
        raise ValueError(
            f"Unknown lado/placement {raw!r}. Use derecha, izquierda, or arriba."
        )
    return PLACEMENT_BY_ALIAS[key]


def map_overlay_kind(raw: str) -> str:
    """Translate Spanish ``tipo`` values to engine kind names.

    Args:
        raw: Cell value (``sticker``, ``etiqueta``, ``enfoque``, …).

    Returns:
        ``sticker``, ``brush_label``, or ``enfoque_lockup``.

    Raises:
        ValueError: Unknown value.
    """
    key = normalise_column_header(raw) or "sticker"
    if key not in KIND_BY_ALIAS:
        raise ValueError(
            f"Unknown tipo {raw!r}. Use sticker, etiqueta, or enfoque."
        )
    return KIND_BY_ALIAS[key]


def overlays_from_rows(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    """Convert Imágenes_y_tiempos rows into YAML overlay mappings.

    Args:
        rows: Dicts from ``02_imagenes_y_tiempos.csv``.

    Returns:
        Overlay list for ``brief.yaml``.

    Raises:
        ValueError: Missing start/end, or unknown tipo/lado.
    """
    overlays: list[dict[str, Any]] = []
    for index, row in enumerate(rows, start=1):
        overlay_id = row.get("id") or row.get("nombre") or f"overlay_{index}"
        start_raw = row.get("tiempo_inicio") or row.get("start") or row.get("desde")
        end_raw = row.get("tiempo_fin") or row.get("end") or row.get("hasta")
        if not start_raw or not end_raw:
            raise ValueError(f"Row {overlay_id}: missing tiempo_inicio / tiempo_fin")
        file_name = row.get("archivo_imagen") or row.get("archivo") or row.get("file")
        text = row.get("texto") or row.get("text")
        notes = (
            row.get("notas_edicion")
            or row.get("notas")
            or row.get("que_es")
            or row.get("que_quieres")
        )
        hint = f"{text or ''} {notes or ''}".lower()
        raw_kind = row.get("tipo") or row.get("kind") or ""
        if raw_kind:
            kind = map_overlay_kind(raw_kind)
        elif "enfoque" in hint:
            kind = "enfoque_lockup"
        elif text and not file_name:
            kind = "brush_label"
        else:
            kind = "sticker"
        side_raw = row.get("lado") or row.get("placement") or ""
        if not side_raw:
            side_raw = "arriba" if kind != "sticker" else "derecha"
        overlay: dict[str, Any] = {
            "id": overlay_id,
            "kind": kind,
            "start": parse_timestamp_to_seconds(start_raw),
            "end": parse_timestamp_to_seconds(end_raw),
            "placement": map_placement_name(side_raw),
        }
        if file_name:
            overlay["file"] = f"overlays/{Path(file_name).name}"
        if text:
            overlay["text"] = text
        if row.get("ancho_max"):
            overlay["max_w"] = int(float(row["ancho_max"]))
        if row.get("alto_max"):
            overlay["max_h"] = int(float(row["alto_max"]))
        if kind == "sticker" and ("max_w" not in overlay or "max_h" not in overlay):
            max_width, max_height, _size_name = sticker_box_for_notes(notes or "")
            overlay.setdefault("max_w", max_width)
            overlay.setdefault("max_h", max_height)
        if row.get("tamano_fuente") or row.get("font_size"):
            overlay["font_size"] = int(
                float(row.get("tamano_fuente") or row["font_size"])
            )
        elif kind == "brush_label":
            overlay["font_size"] = (
                72 if "angélica" in (text or "").lower() or "angelica" in (text or "").lower() else 66
            )
        if notes:
            overlay["notes"] = notes
        overlays.append(overlay)
    return overlays


def speech_corrections_from_rows(rows: list[dict[str, str]]) -> dict[str, str]:
    """Convert the Correcciones tab into the ``asr_fix`` YAML map.

    Args:
        rows: Dicts from ``03_correcciones_transcripcion.csv``.

    Returns:
        Wrong word → correct word.
    """
    corrections: dict[str, str] = {}
    for row in rows:
        wrong = row.get("palabra_incorrecta") or row.get("whisper") or row.get("from")
        right = row.get("palabra_correcta") or row.get("correcto") or row.get("to")
        if wrong and right:
            corrections[wrong] = right
    return corrections


def brief_dict_from_csv_dir(csv_dir: Path) -> dict[str, Any]:
    """Build a brief mapping from a folder of Spanish CSV tabs.

    Args:
        csv_dir: Directory with ``01_proyecto.csv`` and ``02_imagenes_y_tiempos.csv``.

    Returns:
        Canonical brief dict (not yet written to disk).

    Raises:
        FileNotFoundError: Required tabs are missing.
    """
    csv_dir = csv_dir.resolve()
    project_path = find_first_existing_file(
        csv_dir, ["01_proyecto.csv", "proyecto.csv", "project.csv"]
    )
    images_path = find_first_existing_file(
        csv_dir,
        ["02_imagenes_y_tiempos.csv", "imagenes_y_tiempos.csv", "overlays.csv"],
    )
    corrections_path = find_first_existing_file(
        csv_dir,
        ["03_correcciones_transcripcion.csv", "correcciones.csv", "asr_fix.csv"],
    )
    notes_path = find_first_existing_file(
        csv_dir, ["04_notas_edicion.csv", "notas.csv"]
    )
    if project_path is None or images_path is None:
        raise FileNotFoundError(
            f"{csv_dir} needs 01_proyecto.csv and 02_imagenes_y_tiempos.csv"
        )

    project_fields = read_key_value_csv(project_path)
    overlays = overlays_from_rows(read_csv_rows(images_path))
    corrections = (
        speech_corrections_from_rows(read_csv_rows(corrections_path))
        if corrections_path
        else {}
    )
    notes_rows = read_csv_rows(notes_path) if notes_path else []
    notes = [
        f"{row.get('seccion') or 'general'}: {row.get('nota') or row.get('texto') or ''}".strip()
        for row in notes_rows
        if (row.get("nota") or row.get("texto"))
    ]

    slug = project_fields.get("slug") or csv_dir.parent.name
    return {
        "project": {
            "slug": slug,
            "title": project_fields.get("titulo") or project_fields.get("title") or slug,
            "brand": project_fields.get("marca") or project_fields.get("brand") or "dra_angelica",
            "language": project_fields.get("idioma") or "es",
        },
        "source": {
            "video": project_fields.get("video") or "source/talking_head.mp4",
            "transcript": project_fields.get("transcripcion") or "source/transcript.json",
            "overlays": "overlays",
            "trim_start": parse_timestamp_to_seconds(
                project_fields.get("cortar_inicio")
                or project_fields.get("empieza_el_reel")
                or project_fields.get("trim_start")
                or 0
            ),
            "talk_end": parse_timestamp_to_seconds(
                project_fields.get("fin_de_habla") or project_fields.get("talk_end")
            ),
        },
        "output": {
            "filename": project_fields.get("archivo_salida") or f"{slug}.mp4",
            "fade_white": float(project_fields.get("fundido_blanco") or 0.70),
            "endcard_hold": float(project_fields.get("tarjeta_final") or 2.00),
        },
        "captions": {
            "enabled": (project_fields.get("subtitulos") or "si").lower()
            not in ("no", "false", "0"),
            "asr_fix": corrections,
        },
        "overlays": overlays,
        "notes": notes,
    }


def write_brief_from_csv_dir(csv_dir: Path, destination: Path | None = None) -> Path:
    """Convert a CSV folder to brief.yaml next to it (or at ``destination``).

    Args:
        csv_dir: Folder of Spanish CSV tabs.
        destination: Optional ``brief.yaml`` path.

    Returns:
        Path written.
    """
    brief = brief_dict_from_csv_dir(csv_dir)
    if destination is None:
        destination = csv_dir.parent / "brief.yaml"
    write_brief_yaml(brief, destination)
    return destination


def _anchor_row_index(image) -> int | None:
    """Return the 1-based worksheet row an openpyxl image is anchored to.

    Args:
        image: ``openpyxl.drawing.image.Image``.

    Returns:
        Row number, or None if the anchor cannot be read.
    """
    anchor = getattr(image, "anchor", None)
    if anchor is None:
        return None
    if isinstance(anchor, str):
        # "F4" → 4
        digits = "".join(character for character in anchor if character.isdigit())
        return int(digits) if digits else None
    from_cell = getattr(anchor, "_from", None)
    if from_cell is not None and getattr(from_cell, "row", None) is not None:
        return int(from_cell.row) + 1
    return None


def _overlay_sheet_and_header(workbook) -> tuple[Any, int, list[str]] | None:
    """Find the overlay table (simple Brief sheet or legacy multi-tab).

    Args:
        workbook: Open workbook.

    Returns:
        ``(sheet, header_row_1based, normalised_headers)`` or None.
    """
    preferred = []
    for name in ("Brief", "Reel", "Imagenes_y_tiempos"):
        if name in workbook.sheetnames:
            preferred.append(workbook[name])
    preferred.extend(
        sheet
        for sheet in workbook.worksheets
        if sheet not in preferred
        and normalise_column_header(sheet.title)
        not in ("instrucciones", "proyecto", "correcciones", "notas_edicion")
    )
    for sheet in preferred:
        for row_index in range(1, 10):
            headers = [
                normalise_column_header(str(cell.value or ""))
                for cell in sheet[row_index]
            ]
            if "desde" in headers or "tiempo_inicio" in headers:
                return sheet, row_index, headers
    return None


def extract_xlsx_images_to_overlays(
    workbook,
    overlays_directory: Path,
) -> dict[int, str]:
    """Write pictures from the Foto/imagen column into ``overlays/``.

    Rows that share the same image bytes reuse one PNG. Filename comes from
    ``archivo_imagen`` when present, else ``foto_<row>.png``.

    Args:
        workbook: Open workbook (keep ``data_only=False`` so drawings load).
        overlays_directory: Project ``overlays`` folder.

    Returns:
        Excel row number → overlay file name.
    """
    import hashlib

    from modules.video.image_ops import save_pasted_overlay_png

    located = _overlay_sheet_and_header(workbook)
    if located is None:
        return {}
    sheet, header_row, headers = located
    images = getattr(sheet, "_images", None) or []
    if not images:
        return {}

    file_column = None
    for key in ("archivo_imagen", "archivo", "file"):
        if key in headers:
            file_column = headers.index(key) + 1
            break

    overlays_directory.mkdir(parents=True, exist_ok=True)
    digest_to_name: dict[str, str] = {}
    row_to_name: dict[int, str] = {}
    for image in images:
        row_index = _anchor_row_index(image)
        if row_index is None or row_index <= header_row:
            continue
        try:
            payload = image._data()
        except Exception:
            continue
        if not payload:
            continue
        digest = hashlib.sha256(payload).hexdigest()
        if digest in digest_to_name:
            row_to_name[row_index] = digest_to_name[digest]
            continue
        stated_name = ""
        if file_column:
            stated_name = str(sheet.cell(row_index, file_column).value or "").strip()
        file_name = Path(stated_name).name if stated_name else f"foto_{row_index}.png"
        png_name = Path(file_name).with_suffix(".png").name
        try:
            save_pasted_overlay_png(payload, overlays_directory / png_name)
        except ValueError:
            (overlays_directory / png_name).write_bytes(payload)
        digest_to_name[digest] = png_name
        row_to_name[row_index] = png_name
    return row_to_name


def _slugify_title(title: str, fallback: str) -> str:
    """Turn a human title into a folder-safe slug."""
    raw = normalise_column_header(title).replace("—", " ").replace("-", " ")
    parts = [part for part in raw.replace("__", "_").split("_") if part]
    slug = "_".join(parts)[:40]
    return slug or fallback


def brief_dict_from_simple_sheet(
    sheet,
    row_to_filename: dict[int, str],
    project_dir_name: str,
) -> dict[str, Any]:
    """Parse the one-tab Brief sheet (título + Desde/Hasta/Foto/Qué quieres).

    Args:
        sheet: Worksheet named Brief.
        row_to_filename: Pictures extracted from the Foto column.
        project_dir_name: Fallback slug.

    Returns:
        Canonical brief dict.
    """
    located = _overlay_sheet_and_header(sheet.parent)
    if located is None:
        raise ValueError("No Desde/Hasta table on the Brief sheet")
    table_sheet, header_row, headers = located
    title = ""
    trim_raw = "0"
    talk_end_raw = ""
    for row_index in range(1, header_row):
        label = normalise_column_header(str(table_sheet.cell(row_index, 1).value or ""))
        value = str(table_sheet.cell(row_index, 2).value or "").strip()
        if "titulo" in label:
            title = value
        elif (
            "cortar" in label
            or "empieza" in label
            or label.startswith("el_reel")
        ):
            trim_raw = value or "0"
        elif "hablar" in label or "terminas" in label or label.startswith("fin"):
            talk_end_raw = value

    def cell(row_index: int, key: str) -> str:
        if key not in headers:
            return ""
        value = table_sheet.cell(row_index, headers.index(key) + 1).value
        return "" if value is None else str(value).strip()

    overlay_rows: list[dict[str, str]] = []
    for row_index in range(header_row + 1, table_sheet.max_row + 1):
        start_raw = cell(row_index, "desde") or cell(row_index, "tiempo_inicio")
        end_raw = cell(row_index, "hasta") or cell(row_index, "tiempo_fin")
        if not start_raw or not end_raw:
            continue
        file_name = (
            row_to_filename.get(row_index)
            or cell(row_index, "archivo_imagen")
            or cell(row_index, "archivo")
        )
        want = (
            cell(row_index, "que_quieres")
            or cell(row_index, "texto")
            or cell(row_index, "nota")
            or cell(row_index, "notas")
            or cell(row_index, "que_es")
        )
        want_lower = want.lower()
        text = cell(row_index, "texto")
        if not file_name and not text:
            if "enfoque" in want_lower:
                text = "ENFOQUE INTEGRAL"
            elif "angélica" in want_lower or "angelica" in want_lower:
                text = "Dra. Angélica"
            elif "medicina familiar" in want_lower:
                text = "Medicina Familiar"
        side = cell(row_index, "lado")
        if not side:
            if "izquierda" in want_lower:
                side = "izquierda"
            elif "derecha" in want_lower:
                side = "derecha"
            elif "arriba" in want_lower or "cielo" in want_lower:
                side = "arriba"
        overlay_id = cell(row_index, "id") or (
            Path(file_name).stem if file_name else f"item_{len(overlay_rows) + 1}"
        )
        overlay_rows.append(
            {
                "id": overlay_id,
                "tiempo_inicio": start_raw,
                "tiempo_fin": end_raw,
                "archivo_imagen": file_name,
                "texto": text,
                "lado": side,
                "notas_edicion": want,
            }
        )

    if not talk_end_raw:
        raise ValueError("Fill 'Terminas de hablar' at the top of the sheet")
    slug = _slugify_title(title, project_dir_name)
    return {
        "project": {
            "slug": slug,
            "title": title or slug,
            "brand": "dra_angelica",
            "language": "es",
        },
        "source": {
            "video": "source/talking_head.mp4",
            "transcript": "source/transcript.json",
            "overlays": "overlays",
            "trim_start": parse_timestamp_to_seconds(trim_raw or 0),
            "talk_end": parse_timestamp_to_seconds(talk_end_raw),
        },
        "output": {
            "filename": f"{slug}.mp4",
            "fade_white": 0.70,
            "endcard_hold": 2.00,
        },
        "captions": {"enabled": True, "asr_fix": {}},
        "overlays": overlays_from_rows(overlay_rows),
        "notes": [],
    }


def write_brief_from_xlsx(xlsx_path: Path, destination: Path | None = None) -> Path:
    """Convert Angélica's Excel brief (simple one-tab, or legacy multi-tab) to YAML.

    Pictures in the Foto/imagen column are extracted into ``overlays/``.

    Args:
        xlsx_path: Workbook.
        destination: Optional ``brief.yaml`` path.

    Returns:
        Path written.

    Raises:
        ImportError: openpyxl is missing.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to read .xlsx. pip install openpyxl."
        ) from exc

    if destination is None:
        parent = xlsx_path.parent
        destination = (
            parent.parent / "brief.yaml"
            if parent.name == "brief_sheet"
            else parent / "brief.yaml"
        )
    overlays_directory = destination.parent / "overlays"
    drawing_workbook = openpyxl.load_workbook(xlsx_path, data_only=False)
    row_to_filename = extract_xlsx_images_to_overlays(
        drawing_workbook, overlays_directory
    )
    if row_to_filename:
        print(
            f"extracted {len(set(row_to_filename.values()))} overlay image(s) "
            f"→ {overlays_directory}"
        )

    located = _overlay_sheet_and_header(drawing_workbook)
    is_simple = located is not None and "desde" in located[2]
    if is_simple:
        sheet = located[0]
        brief = brief_dict_from_simple_sheet(
            sheet,
            row_to_filename,
            destination.parent.name,
        )
        drawing_workbook.close()
        write_brief_yaml(brief, destination)
        return destination

    drawing_workbook.close()
    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Reuse the CSV importer: dump each known tab, then parse as usual.
    export_dir = xlsx_path.parent / ".sheet_csv_export"
    export_dir.mkdir(parents=True, exist_ok=True)

    name_map = {
        "proyecto": "01_proyecto.csv",
        "project": "01_proyecto.csv",
        "imagenes_y_tiempos": "02_imagenes_y_tiempos.csv",
        "imagenes": "02_imagenes_y_tiempos.csv",
        "overlays": "02_imagenes_y_tiempos.csv",
        "correcciones_transcripcion": "03_correcciones_transcripcion.csv",
        "correcciones": "03_correcciones_transcripcion.csv",
        "notas_edicion": "04_notas_edicion.csv",
        "notas": "04_notas_edicion.csv",
    }
    for sheet in workbook.worksheets:
        key = normalise_column_header(sheet.title)
        if key in ("instrucciones", "como_usar", "readme"):
            continue
        filename = name_map.get(key)
        if filename is None:
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [normalise_column_header(str(header or "")) for header in rows[0]]
        out_path = export_dir / filename
        with out_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for row in rows[1:]:
                if row is None or all(
                    cell is None or str(cell).strip() == "" for cell in row
                ):
                    continue
                writer.writerow(["" if cell is None else str(cell) for cell in row])
    written = write_brief_from_csv_dir(export_dir, destination)
    return written
